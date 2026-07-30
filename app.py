"""Event Guest QR — staff-only guest list, email QR PDFs, scan check-in."""

from __future__ import annotations

import base64
import io
import json
import os
import re
import smtplib
import sqlite3
import ssl
import urllib.error
import urllib.request
import uuid
import zipfile
from datetime import datetime
from email.message import EmailMessage
from functools import wraps
from pathlib import Path

import qrcode
from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "guests.db"
EMAIL_CONFIG_PATH = BASE_DIR / "email_config.json"

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "event-guest-qr-dev-key")
# Needed on Render so CSS/JS URLs use https (avoids "broken"/unstyled pages)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
# Secure cookies on HTTPS hosts (Render). Keep off for local http:// testing.
_public = os.environ.get("PUBLIC_BASE_URL", "")
_on_render = bool(os.environ.get("RENDER"))
app.config["SESSION_COOKIE_SECURE"] = _on_render or _public.startswith("https://")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


# Logins: admin can import Excel; staff can check-in only
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
ADMIN_PASSWORD_HASH = generate_password_hash(ADMIN_PASSWORD)

STAFF_USERNAME = os.environ.get("STAFF_USERNAME", "staff")
STAFF_PASSWORD = os.environ.get("STAFF_PASSWORD", "event123")
STAFF_PASSWORD_HASH = generate_password_hash(STAFF_PASSWORD)

EVENT_NAME = "NGPA 2026 Annual Conference"
EVENT_DATE = "8 & 9 August 2026 · Saturday & Sunday"
EVENT_VENUE = "North Gujarat Physician Association"

# Public base URL embedded inside QR codes (must be reachable by staff phones).
# Example: http://10.197.190.212:5050
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")

# Guest categories — badge background art in static/badges/
CATEGORIES = ("Delegate", "Faculty", "Organizer", "Pharma", "Guest")
CATEGORY_COLORS = {
    # RGB 0-1 for UI accents (matches badge footer colors)
    "Delegate": {"rgb": (0.10, 0.25, 0.55), "css": "cat-delegate"},  # blue
    "Faculty": {"rgb": (0.75, 0.12, 0.14), "css": "cat-faculty"},  # red
    "Organizer": {"rgb": (0.90, 0.45, 0.08), "css": "cat-organizer"},  # orange
    "Pharma": {"rgb": (0.15, 0.55, 0.28), "css": "cat-pharma"},  # green
    "Guest": {"rgb": (0.45, 0.18, 0.65), "css": "cat-guest"},  # purple
}
# Prefer vector PDFs in category/ (sharp on phone). PNG under static/badges/ is fallback only.
CATEGORY_BADGE_PDF = {
    "Delegate": "DELEGATE.pdf",
    "Faculty": "FACULTY.pdf",
    "Organizer": "ORGANIZER.pdf",
    "Pharma": "PHARMA.pdf",
    "Guest": "GUEST.pdf",
}
CATEGORY_BADGE_BG = {
    "Delegate": "delegate.png",
    "Faculty": "faculty.png",
    "Organizer": "organizer.png",
    "Pharma": "pharma.png",
    "Guest": "guest.png",
}
CATEGORY_ALIASES = {
    "delegate": "Delegate",
    "delicate": "Delegate",  # common typo
    "faculty": "Faculty",
    "organizer": "Organizer",
    "organiser": "Organizer",
    "origanizer": "Organizer",  # typo
    "pharma": "Pharma",
    "pharmaceutical": "Pharma",
    "guest": "Guest",
}


def normalize_category(value: str | None) -> str:
    raw = (value or "").strip()
    if not raw:
        return "Delegate"
    mapped = CATEGORY_ALIASES.get(raw.lower())
    if mapped:
        return mapped
    # Title-case exact match against known list
    for cat in CATEGORIES:
        if raw.lower() == cat.lower():
            return cat
    return "Delegate"


SAMPLE_GUESTS = [
    # Delegate
    {
        "name": "Shruti Anandas",
        "email": "shrutianandas123@gmail.com",
        "phone": "9326305627",
        "specialty": "General Medicine",
        "city": "Ahmedabad",
        "designation": "Delegate",
        "meal": "Vegetarian",
    },
    {
        "name": "Vishal Patel",
        "email": "vickyashokpatel123@gmail.com",
        "phone": "8866325109",
        "specialty": "Critical Care Medicine",
        "city": "Mehsana",
        "designation": "Delegate",
        "meal": "Non-Vegetarian",
    },
    {
        "name": "Aarav Sharma",
        "email": "aarav.sharma.test@example.com",
        "phone": "9876543210",
        "specialty": "Cardiology",
        "city": "Surat",
        "designation": "Delegate",
        "meal": "Vegetarian",
    },
    # Faculty
    {
        "name": "Rohan Mehta",
        "email": "rohan.mehta.test@example.com",
        "phone": "9988776655",
        "specialty": "Orthopedics",
        "city": "Rajkot",
        "designation": "Faculty",
        "meal": "Vegan",
    },
    {
        "name": "Dr. Neha Kapoor",
        "email": "neha.kapoor.test@example.com",
        "phone": "9900112233",
        "specialty": "Cardiology",
        "city": "Ahmedabad",
        "designation": "Faculty",
        "meal": "Vegetarian",
    },
    {
        "name": "Dr. Sameer Joshi",
        "email": "sameer.joshi.test@example.com",
        "phone": "9911223344",
        "specialty": "Neurology",
        "city": "Vadodara",
        "designation": "Faculty",
        "meal": "Non-Vegetarian",
    },
    # Organizer
    {
        "name": "Ananya Gupta",
        "email": "ananya.gupta.test@example.com",
        "phone": "9811122233",
        "specialty": "Event Ops",
        "city": "Ahmedabad",
        "designation": "Organizer",
        "meal": "Vegetarian",
    },
    {
        "name": "Karan Desai",
        "email": "karan.desai.test@example.com",
        "phone": "9822233344",
        "specialty": "Registration Desk",
        "city": "Ahmedabad",
        "designation": "Organizer",
        "meal": "Non-Vegetarian",
    },
    {
        "name": "Pooja Shah",
        "email": "pooja.shah.test@example.com",
        "phone": "9833344455",
        "specialty": "Hospitality",
        "city": "Gandhinagar",
        "designation": "Organizer",
        "meal": "Vegetarian",
    },
    # Pharma
    {
        "name": "Meera Nair",
        "email": "meera.nair.test@example.com",
        "phone": "9844455566",
        "specialty": "Medical Affairs",
        "city": "Mumbai",
        "designation": "Pharma",
        "meal": "Vegan",
    },
    {
        "name": "Rahul Verma",
        "email": "rahul.verma.test@example.com",
        "phone": "9855566677",
        "specialty": "Sales",
        "city": "Pune",
        "designation": "Pharma",
        "meal": "Non-Vegetarian",
    },
    {
        "name": "Isha Trivedi",
        "email": "isha.trivedi.test@example.com",
        "phone": "9866677788",
        "specialty": "Marketing",
        "city": "Ahmedabad",
        "designation": "Pharma",
        "meal": "Vegetarian",
    },
]

SAMPLE_XLSX = BASE_DIR / "DELEGATES.xlsx"
if not SAMPLE_XLSX.exists():
    SAMPLE_XLSX = BASE_DIR / "guests_categories_sample.xlsx"
if not SAMPLE_XLSX.exists():
    SAMPLE_XLSX = BASE_DIR / "guests_sample.xlsx"

# Fixed event guest lists (category Excel files in project folder)
FIXED_GUEST_EXCELS = (
    (BASE_DIR / "DELEGATES.xlsx", "Delegate"),
    (BASE_DIR / "PHARMA.xlsx", "Pharma"),
    (BASE_DIR / "ORGANISER.xlsx", "Organizer"),
)


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, decl: str) -> None:
    cols = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})")}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl}")


def init_db() -> None:
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT NOT NULL,
            meal TEXT NOT NULL,
            table_no TEXT NOT NULL,
            token TEXT NOT NULL UNIQUE,
            arrived INTEGER NOT NULL DEFAULT 0,
            arrived_at TEXT,
            lunch_claimed INTEGER NOT NULL DEFAULT 0,
            claimed_at TEXT,
            pdf_sent INTEGER NOT NULL DEFAULT 0,
            pdf_sent_at TEXT
        )
        """
    )
    _ensure_column(conn, "guests", "arrived", "INTEGER NOT NULL DEFAULT 0")
    _ensure_column(conn, "guests", "arrived_at", "TEXT")
    _ensure_column(conn, "guests", "pdf_sent", "INTEGER NOT NULL DEFAULT 0")
    _ensure_column(conn, "guests", "pdf_sent_at", "TEXT")
    _ensure_column(conn, "guests", "dinner_claimed", "INTEGER NOT NULL DEFAULT 0")
    _ensure_column(conn, "guests", "dinner_claimed_at", "TEXT")
    _ensure_column(conn, "guests", "email_acked", "INTEGER NOT NULL DEFAULT 0")
    _ensure_column(conn, "guests", "email_acked_at", "TEXT")
    _ensure_column(conn, "guests", "specialty", "TEXT DEFAULT ''")
    _ensure_column(conn, "guests", "city", "TEXT DEFAULT ''")
    _ensure_column(conn, "guests", "designation", "TEXT DEFAULT ''")

    conn.commit()
    conn.close()


def _norm_email(email: str) -> str:
    return (email or "").strip().lower()


def _norm_name(name: str) -> str:
    return " ".join((name or "").strip().lower().split())


def _norm_phone(phone: str) -> str:
    digits = re.sub(r"\D+", "", phone or "")
    # Prefer last 10 digits (common for Indian mobiles with +91 / 0 prefix)
    if len(digits) > 10:
        return digits[-10:]
    return digits


def _guest_identity_key(name: str, phone: str) -> str:
    return f"{_norm_name(name)}|{_norm_phone(phone)}"


def _guest_insert_values(guest: dict, token: str | None = None) -> tuple:
    return (
        guest["name"],
        (guest.get("email") or "").strip(),
        guest.get("phone") or "",
        guest.get("meal") or "Vegetarian",
        guest.get("table_no") or "",
        guest.get("specialty") or "",
        guest.get("city") or "",
        normalize_category(guest.get("designation") or guest.get("category")),
        token or uuid.uuid4().hex,
    )


def _insert_sample_guests(conn: sqlite3.Connection) -> None:
    for guest in SAMPLE_GUESTS:
        conn.execute(
            """
            INSERT INTO guests
            (name, email, phone, meal, table_no, specialty, city, designation, token)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            _guest_insert_values(guest),
        )


def sync_guests_from_rows(
    rows: list[dict],
    *,
    replace_all: bool = False,
    category_scope: str | None = None,
) -> dict:
    """
    Merge Excel into the DB without changing QR/token for returning guests.

    Same name + phone = same person → keep token, short ID, and meal check-in flags.
    New name/phone pair → new token.

    replace_all: wipe every guest first, then import.
    category_scope: only remove missing guests in that category (other categories kept).
    """
    conn = get_db()
    if replace_all:
        conn.execute("DELETE FROM guests")
        existing = []
    else:
        existing = conn.execute("SELECT * FROM guests").fetchall()

    by_identity = {
        _guest_identity_key(row["name"], row["phone"]): row for row in existing
    }
    scope = normalize_category(category_scope) if category_scope else None

    seen: set[str] = set()
    kept = 0
    created = 0

    for guest in rows:
        if not _norm_name(guest["name"]) or not _norm_phone(guest.get("phone") or ""):
            continue
        key = _guest_identity_key(guest["name"], guest.get("phone") or "")
        if key in seen:
            continue
        seen.add(key)

        designation = normalize_category(
            guest.get("designation") or guest.get("category") or scope or "Delegate"
        )
        old = by_identity.get(key)
        email_value = (guest.get("email") or "").strip()
        if old:
            if not email_value:
                email_value = (old["email"] or "").strip()
            conn.execute(
                """
                UPDATE guests
                SET name = ?, email = ?, phone = ?, meal = ?, table_no = ?,
                    specialty = ?, city = ?, designation = ?
                WHERE id = ?
                """,
                (
                    guest["name"].strip(),
                    email_value,
                    guest.get("phone") or "",
                    guest.get("meal") or "Vegetarian",
                    guest.get("table_no") or "",
                    guest.get("specialty") or "",
                    guest.get("city") or "",
                    designation,
                    old["id"],
                ),
            )
            kept += 1
        else:
            payload = dict(guest)
            payload["designation"] = designation
            conn.execute(
                """
                INSERT INTO guests
                (name, email, phone, meal, table_no, specialty, city, designation, token)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                _guest_insert_values(payload),
            )
            created += 1

    removed = 0
    for key, old in by_identity.items():
        if key in seen:
            continue
        if scope and normalize_category(old["designation"]) != scope:
            continue
        if not replace_all and scope is None:
            # Full-file sync: remove anyone missing from the file
            conn.execute("DELETE FROM guests WHERE id = ?", (old["id"],))
            removed += 1
        elif scope:
            conn.execute("DELETE FROM guests WHERE id = ?", (old["id"],))
            removed += 1

    conn.commit()
    total = conn.execute("SELECT COUNT(*) AS c FROM guests").fetchone()["c"]
    conn.close()
    return {
        "total": total,
        "kept": kept,
        "created": created,
        "removed": removed,
    }


def replace_guests_from_rows(rows: list[dict]) -> int:
    """Backward-compatible wrapper; prefer sync_guests_from_rows."""
    return sync_guests_from_rows(rows)["total"]


def read_guests_from_excel(
    file_storage, default_category: str | None = None
) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise RuntimeError("Excel file is empty.")

    normalized = [str(h or "").strip().lower() for h in header]
    aliases = {
        "name": {"name", "guest", "guest name", "full name"},
        "first_name": {"first name", "firstname", "first"},
        "last_name": {"last name", "lastname", "last", "surname"},
        "email": {"email", "email id", "mail"},
        "phone": {
            "phone",
            "mobile",
            "contact",
            "phone number",
            "mobile number",
            "mob",
            "mobile no",
            "mobile no.",
            "number",
            "num",
        },
        "specialty": {"specialty", "speciality", "field", "department"},
        "city": {"city", "place", "location", "company"},
        "designation": {
            "designation",
            "role",
            "category",
            "type",
            "group",
            "badge",
        },
        "meal": {"meal", "meal preference", "preference", "food"},
        "sr_no": {
            "sr no",
            "sr. no",
            "sr.no",
            "srno",
            "s.no",
            "serial",
            "serial no",
            "no",
        },
    }

    def find_col(keys: set[str]) -> int | None:
        for i, name in enumerate(normalized):
            if name in keys:
                return i
        return None

    idx = {key: find_col(vals) for key, vals in aliases.items()}
    has_full_name = idx["name"] is not None
    has_split_name = idx["first_name"] is not None or idx["last_name"] is not None
    if idx["phone"] is None or (not has_full_name and not has_split_name):
        raise RuntimeError(
            "Excel must have NAME and MOBILE NUMBER columns "
            "(optional: CITY, CATEGORY / or pick category in the form)."
        )

    def cell(row, key: str, default: str = "") -> str:
        i = idx[key]
        if i is None:
            return default
        raw = row[i]
        if raw is None:
            return default
        # Keep mobile numbers as digits (Excel may store as float)
        if key == "phone" and isinstance(raw, (int, float)):
            return str(int(raw))
        return str(raw).strip() or default

    fallback_cat = normalize_category(default_category or "Delegate")
    guests: list[dict] = []
    for row in rows_iter:
        if not row:
            continue
        name = cell(row, "name")
        if not name:
            name = " ".join(
                part
                for part in (cell(row, "first_name"), cell(row, "last_name"))
                if part
            ).strip()
        phone = cell(row, "phone")
        email = cell(row, "email")
        if not name and not phone:
            continue
        if not name or not _norm_phone(phone):
            continue
        if email and "@" not in email:
            email = ""
        cat_cell = cell(row, "designation")
        guests.append(
            {
                "name": name,
                "email": email,
                "phone": phone,
                "specialty": cell(row, "specialty"),
                "city": cell(row, "city"),
                "designation": normalize_category(cat_cell) if cat_cell else fallback_cat,
                "meal": cell(row, "meal", "Vegetarian"),
                "table_no": "",
            }
        )
    wb.close()
    if not guests:
        raise RuntimeError(
            "No valid guest rows found. Each row needs NAME and MOBILE NUMBER."
        )
    if len(guests) > 2000:
        raise RuntimeError("Too many rows (max 2000). Split the file.")
    return guests


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        if session.get("role") != "admin":
            flash("Only admin can import guest data.", "bad")
            return redirect(url_for("home"))
        return view(*args, **kwargs)

    return wrapped


def get_guest_by_token(token: str) -> sqlite3.Row | None:
    conn = get_db()
    guest = conn.execute(
        "SELECT * FROM guests WHERE token = ?", (token,)
    ).fetchone()
    conn.close()
    return guest


def get_guest_by_id(guest_id: int) -> sqlite3.Row | None:
    conn = get_db()
    guest = conn.execute(
        "SELECT * FROM guests WHERE id = ?", (guest_id,)
    ).fetchone()
    conn.close()
    return guest


def guest_badge_id(token: str) -> str:
    """Printed badge ID — first 8 chars of token (e.g. A1B2C3D4)."""
    return ((token or "")[:8] or "GUEST").upper()


def resolve_guest_code(raw: str) -> sqlite3.Row | None:
    """Resolve full token, check-in URL, or short badge ID (first 8 of token)."""
    value = (raw or "").strip()
    if not value:
        return None

    match = re.search(r"/checkin/([a-fA-F0-9]{32})", value)
    if match:
        return get_guest_by_token(match.group(1).lower())

    code = value.replace(" ", "").lower()
    if code.startswith("ngpa_"):
        code = code[5:]
    if re.fullmatch(r"[a-f0-9]{32}", code):
        return get_guest_by_token(code)

    # Short ID printed on PDF (e.g. 9168B65B)
    if re.fullmatch(r"[a-f0-9]{6,12}", code):
        conn = get_db()
        rows = conn.execute(
            "SELECT * FROM guests WHERE lower(substr(token, 1, ?)) = ?",
            (len(code), code),
        ).fetchall()
        conn.close()
        if len(rows) == 1:
            return rows[0]
        return None

    return None


def public_url(endpoint: str, **values) -> str:
    path = url_for(endpoint, **values)
    if PUBLIC_BASE_URL:
        return f"{PUBLIC_BASE_URL}{path}"
    return request.url_root.rstrip("/") + path


def guest_qr_url(token: str) -> str:
    """One QR / one guest ID — used for both lunch and dinner check-in."""
    return public_url("checkin", token=token)


def make_qr_image(data: str):
    """Black QR modules on a transparent background (PNG RGBA)."""
    qr = qrcode.QRCode(version=1, box_size=8, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    matrix = qr.get_matrix()
    scale = 8
    size = len(matrix) * scale
    from PIL import Image as PILImage

    img = PILImage.new("RGBA", (size, size), (0, 0, 0, 0))
    pixels = img.load()
    for y, row in enumerate(matrix):
        for x, dark in enumerate(row):
            if not dark:
                continue
            x0, y0 = x * scale, y * scale
            for dy in range(scale):
                for dx in range(scale):
                    pixels[x0 + dx, y0 + dy] = (0, 0, 0, 255)
    return img


def _draw_qr_on_pdf(pdf: canvas.Canvas, data: str, x: float, y: float, size: float) -> None:
    """Draw QR as black modules only — no white fill (badge art shows through)."""
    qr = qrcode.QRCode(version=1, box_size=1, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    matrix = qr.get_matrix()
    n = len(matrix)
    cell = size / n
    pdf.setFillColorRGB(0, 0, 0)
    for row_i, row in enumerate(matrix):
        for col_i, dark in enumerate(row):
            if dark:
                # Matrix row 0 is top; PDF y grows upward
                pdf.rect(
                    x + col_i * cell,
                    y + (n - 1 - row_i) * cell,
                    cell + 0.1,
                    cell + 0.1,
                    fill=1,
                    stroke=0,
                )


def _guest_field(guest: sqlite3.Row, key: str, default: str = "") -> str:
    try:
        value = guest[key]
    except (KeyError, IndexError):
        return default
    return (str(value).strip() if value is not None else "") or default


def category_style(guest: sqlite3.Row | dict) -> dict:
    if isinstance(guest, dict):
        cat = normalize_category(guest.get("designation"))
    else:
        cat = normalize_category(_guest_field(guest, "designation", "Delegate"))
    return {"name": cat, **CATEGORY_COLORS[cat]}


def _badge_background_pdf_path(category: str) -> Path | None:
    """Vector category artwork (category/*.pdf) — keeps header text sharp on phones."""
    filename = CATEGORY_BADGE_PDF.get(category) or CATEGORY_BADGE_PDF["Delegate"]
    path = BASE_DIR / "category" / filename
    return path if path.exists() else None


def _badge_background_path(category: str) -> Path | None:
    """Raster fallback (static/badges/*.png) if category PDF is missing."""
    filename = CATEGORY_BADGE_BG.get(category) or CATEGORY_BADGE_BG["Delegate"]
    path = BASE_DIR / "static" / "badges" / filename
    return path if path.exists() else None


def _merge_vector_badge_background(
    overlay_pdf: io.BytesIO, bg_pdf_path: Path, badge_w: float, badge_h: float
) -> io.BytesIO:
    """Place category PDF under the ReportLab overlay (vector, not rasterized)."""
    import fitz  # PyMuPDF

    out = fitz.open()
    page = out.new_page(width=badge_w, height=badge_h)
    bg = fitz.open(bg_pdf_path)
    try:
        page.show_pdf_page(page.rect, bg, 0)
    finally:
        bg.close()
    overlay = fitz.open(stream=overlay_pdf.getvalue(), filetype="pdf")
    try:
        page.show_pdf_page(page.rect, overlay, 0)
    finally:
        overlay.close()
    buffer = io.BytesIO(out.tobytes(deflate=True, garbage=3))
    out.close()
    buffer.seek(0)
    return buffer


def build_ticket_pdf(guest: sqlite3.Row) -> io.BytesIO:
    """NGPA badge PDF: category background + name, city, QR in the blank middle."""
    # Matches provided artwork aspect (~670×1024)
    badge_w, badge_h = 105 * mm, 160.5 * mm
    style = category_style(guest)
    accent = style["rgb"]
    ink = (0.12, 0.12, 0.16)

    # Prefer vector category/*.pdf (sharp headers). PNG is a soft fallback only.
    bg_pdf_path = _badge_background_pdf_path(style["name"])
    bg_png_path = _badge_background_path(style["name"]) if not bg_pdf_path else None

    def _paint_overlay(pdf: canvas.Canvas, *, draw_raster_bg: bool) -> None:
        if draw_raster_bg and bg_png_path:
            pdf.drawImage(
                str(bg_png_path),
                0,
                0,
                width=badge_w,
                height=badge_h,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )
        elif draw_raster_bg and not bg_pdf_path:
            pdf.setFillColorRGB(1, 1, 1)
            pdf.rect(0, 0, badge_w, badge_h, fill=1, stroke=0)
            pdf.setFillColorRGB(*accent)
            pdf.rect(0, 0, badge_w, 18 * mm, fill=1, stroke=0)
            pdf.setFillColorRGB(1, 1, 1)
            pdf.setFont("Helvetica-Bold", 11)
            pdf.drawCentredString(badge_w / 2, 6.5 * mm, style["name"].upper())

        name = _guest_field(guest, "name").upper()
        city = _guest_field(guest, "city").upper()
        code = guest_badge_id(_guest_field(guest, "token"))

        center_x = badge_w / 2
        name_y = badge_h - 68 * mm
        city_y = badge_h - 76 * mm
        qr_size = 34 * mm
        qr_y = 46 * mm
        qr_x = (badge_w - qr_size) / 2

        pdf.setFillColorRGB(*ink)
        pdf.setFont("Helvetica-Bold", 13)
        max_chars = 24
        if len(name) <= max_chars:
            pdf.drawCentredString(center_x, name_y, name)
        else:
            pdf.drawCentredString(center_x, name_y + 3.5 * mm, name[:max_chars])
            pdf.drawCentredString(
                center_x, name_y - 2.5 * mm, name[max_chars : max_chars * 2]
            )
            city_y = badge_h - 80 * mm

        if city:
            pdf.setFont("Helvetica", 10)
            pdf.drawCentredString(center_x, city_y, city)

        _draw_qr_on_pdf(pdf, guest_qr_url(guest["token"]), qr_x, qr_y, qr_size)

        qr_id_gap = 3.6 * mm
        pdf.setFillColorRGB(*ink)
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawCentredString(center_x, qr_y - qr_id_gap, code)

    # Overlay only (transparent) when merging vector PDF underneath.
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(badge_w, badge_h))
    _paint_overlay(pdf, draw_raster_bg=not bool(bg_pdf_path))
    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    if bg_pdf_path:
        try:
            return _merge_vector_badge_background(buffer, bg_pdf_path, badge_w, badge_h)
        except Exception:
            # PyMuPDF missing/failed → rebuild with PNG raster background
            bg_png_path = _badge_background_path(style["name"])
            buffer = io.BytesIO()
            pdf = canvas.Canvas(buffer, pagesize=(badge_w, badge_h))
            _paint_overlay(pdf, draw_raster_bg=True)
            pdf.showPage()
            pdf.save()
            buffer.seek(0)
    return buffer


def load_email_config() -> dict | None:
    """Prefer env vars (Render), fall back to local email_config.json."""
    resend_key = os.environ.get("RESEND_API_KEY", "").strip()
    if resend_key:
        return {
            "provider": "resend",
            "api_key": resend_key,
            "from_email": os.environ.get(
                "SMTP_FROM_EMAIL", "onboarding@resend.dev"
            ).strip(),
            "from_name": os.environ.get("SMTP_FROM_NAME", "Event Desk").strip(),
        }

    env_user = os.environ.get("SMTP_USER", "").strip()
    env_pass = os.environ.get("SMTP_PASSWORD", "").strip().replace(" ", "")
    if env_user and env_pass:
        port_raw = os.environ.get("SMTP_PORT", "587").strip() or "587"
        try:
            port = int(port_raw)
        except ValueError as exc:
            raise RuntimeError(f"SMTP_PORT must be a number, got {port_raw!r}") from exc
        return {
            "provider": "smtp",
            "smtp_host": os.environ.get("SMTP_HOST", "smtp.gmail.com").strip(),
            "smtp_port": port,
            "smtp_user": env_user,
            "smtp_password": env_pass,
            "from_email": os.environ.get("SMTP_FROM_EMAIL", env_user).strip(),
            "from_name": os.environ.get("SMTP_FROM_NAME", "Event Desk").strip(),
        }
    if EMAIL_CONFIG_PATH.exists():
        with EMAIL_CONFIG_PATH.open(encoding="utf-8") as f:
            data = json.load(f)
        data["provider"] = data.get("provider", "smtp")
        if data.get("smtp_password"):
            data["smtp_password"] = str(data["smtp_password"]).replace(" ", "")
        return data
    return None


def email_is_configured() -> bool:
    try:
        return load_email_config() is not None
    except Exception:  # noqa: BLE001
        return False


def _ack_url(guest: sqlite3.Row) -> str:
    return public_url("acknowledge", token=guest["token"])


def _email_body_text(guest: sqlite3.Row) -> str:
    ack = _ack_url(guest)
    return (
        f"Hi {guest['name']},\n\n"
        f"Please find your badge PDF attached for {EVENT_NAME}.\n"
        f"{EVENT_DATE}\n{EVENT_VENUE}\n\n"
        "Your badge has one QR (guest ID). "
        "Show the same QR at lunch and dinner counters.\n\n"
        "Please confirm you received this pass by opening this link:\n"
        f"{ack}\n\n"
        "See you there!\n"
    )


def _email_body_html(guest: sqlite3.Row) -> str:
    ack = _ack_url(guest)
    return f"""
    <p>Hi {guest['name']},</p>
    <p>Please find your badge PDF attached for <strong>{EVENT_NAME}</strong>.</p>
    <p>{EVENT_DATE}<br/>{EVENT_VENUE}</p>
    <p>Your badge has <strong>one QR</strong> (guest ID).
       Show the same QR at lunch and dinner counters.</p>
    <p>
      <a href="{ack}"
         style="display:inline-block;background:#5b2d8e;color:#fff;padding:12px 18px;
                border-radius:8px;text-decoration:none;font-weight:600;">
        I received my QR pass
      </a>
    </p>
    <p style="color:#666;font-size:13px;">If the button does not work, open:<br/>{ack}</p>
    """


def _send_via_resend(cfg: dict, guest: sqlite3.Row, pdf_bytes: bytes, filename: str) -> None:
    payload = {
        "from": f"{cfg['from_name']} <{cfg['from_email']}>",
        "to": [guest["email"]],
        "subject": f"Your QR pass - {EVENT_NAME}",
        "text": _email_body_text(guest),
        "html": _email_body_html(guest),
        "attachments": [
            {
                "filename": filename,
                "content": base64.b64encode(pdf_bytes).decode("ascii"),
            }
        ],
    }
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {cfg['api_key']}",
            "Content-Type": "application/json",
            # Required by Resend — missing User-Agent causes 403 error 1010
            "User-Agent": "EventGuestQR/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            if resp.status >= 400:
                raise RuntimeError(f"Resend HTTP {resp.status}")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:300]
        raise RuntimeError(f"Resend error {exc.code}: {detail}") from exc


def _smtp_send_message(
    host: str, port: int, user: str, password: str, msg: EmailMessage
) -> None:
    context = ssl.create_default_context()
    if port == 465:
        with smtplib.SMTP_SSL(host, port, timeout=30, context=context) as server:
            server.login(user, password)
            server.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=30) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
            server.login(user, password)
            server.send_message(msg)


def _send_via_smtp(cfg: dict, guest: sqlite3.Row, pdf_bytes: bytes, filename: str) -> None:
    msg = EmailMessage()
    msg["Subject"] = f"Your QR pass - {EVENT_NAME}"
    msg["From"] = f"{cfg.get('from_name', 'Event Desk')} <{cfg['from_email']}>"
    msg["To"] = guest["email"]
    msg.set_content(_email_body_text(guest))
    msg.add_alternative(_email_body_html(guest), subtype="html")
    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=filename,
    )

    host = cfg["smtp_host"]
    port = int(cfg["smtp_port"])
    user = cfg["smtp_user"]
    password = str(cfg["smtp_password"]).replace(" ", "")

    # Try configured port, then the other common Gmail port
    ports_to_try = [port]
    for alt in (465, 587):
        if alt not in ports_to_try:
            ports_to_try.append(alt)

    last_error: Exception | None = None
    for try_port in ports_to_try:
        try:
            _smtp_send_message(host, try_port, user, password, msg)
            return
        except smtplib.SMTPAuthenticationError as exc:
            raise RuntimeError(
                "Gmail login failed. Use a 16-character Google App Password "
                "(Google Account -> Security -> 2-Step Verification -> App passwords). "
                "Do not use your normal Gmail password."
            ) from exc
        except (smtplib.SMTPConnectError, TimeoutError, OSError) as exc:
            last_error = exc
            continue

    raise RuntimeError(
        "Could not connect to Gmail SMTP (ports 587/465 blocked or network issue). "
        "Try: turn OFF office VPN, use home/mobile hotspot, then retry. "
        f"Technical: {last_error}"
    ) from last_error


def send_pdf_email(guest: sqlite3.Row, pdf_buffer: io.BytesIO) -> None:
    cfg = load_email_config()
    if not cfg:
        raise RuntimeError(
            "Email not configured. Set SMTP_USER + SMTP_PASSWORD, or RESEND_API_KEY."
        )

    pdf_bytes = pdf_buffer.getvalue()
    filename = f"{guest['name'].replace(' ', '_')}_pass.pdf"
    provider = cfg.get("provider", "smtp")

    if provider == "resend":
        _send_via_resend(cfg, guest, pdf_bytes, filename)
    else:
        _send_via_smtp(cfg, guest, pdf_bytes, filename)


def mark_pdf_sent(guest_id: int) -> None:
    conn = get_db()
    conn.execute(
        "UPDATE guests SET pdf_sent = 1, pdf_sent_at = ? WHERE id = ?",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), guest_id),
    )
    conn.commit()
    conn.close()


@app.context_processor
def inject_event():
    role = session.get("role")
    return {
        "event_name": EVENT_NAME,
        "event_date": EVENT_DATE,
        "event_venue": EVENT_VENUE,
        "email_ready": email_is_configured(),
        "staff_logged_in": bool(session.get("user")),
        "is_admin": role == "admin",
        "user_role": role,
        "categories": CATEGORIES,
        "category_colors": CATEGORY_COLORS,
        "normalize_category": normalize_category,
        "category_style": category_style,
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"):
        return redirect(url_for("home"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and check_password_hash(
            ADMIN_PASSWORD_HASH, password
        ):
            session["user"] = True
            session["role"] = "admin"
            nxt = request.args.get("next") or url_for("home")
            return redirect(nxt)
        if username == STAFF_USERNAME and check_password_hash(
            STAFF_PASSWORD_HASH, password
        ):
            session["user"] = True
            session["role"] = "staff"
            nxt = request.args.get("next") or url_for("home")
            return redirect(nxt)
        error = "Invalid username or password."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def home():
    q = (request.args.get("q") or "").strip()
    raw_cat = (request.args.get("category") or "").strip()
    category = (
        normalize_category(raw_cat)
        if raw_cat and raw_cat.lower() != "all"
        else ""
    )
    status = (request.args.get("status") or "").strip().lower()
    if status not in {"", "all", "arrived", "lunch", "dinner"}:
        status = ""
    if status == "all":
        status = ""

    conn = get_db()
    all_guests = conn.execute("SELECT * FROM guests ORDER BY id").fetchall()
    category_counts = {
        cat: sum(1 for g in all_guests if normalize_category(g["designation"]) == cat)
        for cat in CATEGORIES
    }

    # Counts + list scoped to selected category (if any)
    scoped = list(all_guests)
    if category:
        scoped = [
            g for g in scoped if normalize_category(g["designation"]) == category
        ]

    arrived = sum(1 for g in scoped if g["arrived"])
    lunch_claimed = sum(1 for g in scoped if g["lunch_claimed"])
    dinner_claimed = sum(1 for g in scoped if g["dinner_claimed"])

    guests = scoped
    if status == "arrived":
        guests = [g for g in guests if g["arrived"]]
    elif status == "lunch":
        guests = [g for g in guests if g["lunch_claimed"]]
    elif status == "dinner":
        guests = [g for g in guests if g["dinner_claimed"]]

    if q:
        needle = q.lower()
        guests = [
            g
            for g in guests
            if needle in (g["name"] or "").lower()
            or needle in (g["email"] or "").lower()
            or needle in (g["phone"] or "").lower()
            or needle in (g["specialty"] or "").lower()
            or needle in (g["city"] or "").lower()
            or needle in (g["designation"] or "").lower()
            or needle in (g["meal"] or "").lower()
        ]
    conn.close()

    return render_template(
        "home.html",
        guests=guests,
        arrived=arrived,
        lunch_claimed=lunch_claimed,
        dinner_claimed=dinner_claimed,
        total=len(all_guests),
        scoped_total=len(scoped),
        shown=len(guests),
        q=q,
        category=category,
        status=status,
        category_counts=category_counts,
    )


@app.route("/upload-excel", methods=["POST"])
@admin_required
def upload_excel():
    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Please choose an Excel (.xlsx) file.", "bad")
        return redirect(url_for("home"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "bad")
        return redirect(url_for("home"))
    category = normalize_category(request.form.get("import_category") or "Delegate")
    replace_all = request.form.get("replace_all") == "1"
    try:
        rows = read_guests_from_excel(file, default_category=category)
        stats = sync_guests_from_rows(
            rows,
            replace_all=replace_all,
            category_scope=None if replace_all else category,
        )
        mode = "replaced all" if replace_all else f"merged into {category}"
        flash(
            f"Imported {stats['total']} guest(s) ({mode}): "
            f"{stats['kept']} kept same QR/ID, "
            f"{stats['created']} new, "
            f"{stats['removed']} removed.",
            "ok",
        )
    except Exception as exc:  # noqa: BLE001
        flash(f"Excel upload failed: {exc}", "bad")
    return redirect(url_for("home"))


@app.route("/clear-guests", methods=["POST"])
@admin_required
def clear_guests():
    scope = (request.form.get("clear_category") or "").strip()
    conn = get_db()
    if scope and scope.lower() != "all":
        cat = normalize_category(scope)
        count = conn.execute(
            "SELECT COUNT(*) AS c FROM guests WHERE designation = ?",
            (cat,),
        ).fetchone()["c"]
        conn.execute("DELETE FROM guests WHERE designation = ?", (cat,))
        conn.commit()
        conn.close()
        flash(f"Cleared {count} {cat} guest(s). Other categories kept.", "ok")
    else:
        count = conn.execute("SELECT COUNT(*) AS c FROM guests").fetchone()["c"]
        conn.execute("DELETE FROM guests")
        conn.commit()
        conn.close()
        flash(f"Cleared {count} guest(s) (all categories).", "ok")
    return redirect(url_for("home"))


def load_fixed_guest_excels(*, replace_all: bool = True) -> dict:
    """Load DELEGATES / PHARMA / ORGANISER Excel files from the project folder."""
    all_rows: list[dict] = []
    loaded_files: list[str] = []
    for path, category in FIXED_GUEST_EXCELS:
        if not path.exists():
            continue
        rows = read_guests_from_excel(path, default_category=category)
        all_rows.extend(rows)
        loaded_files.append(f"{path.name} ({len(rows)} {category})")
    if not all_rows:
        raise RuntimeError(
            "No fixed Excel files found. Add DELEGATES.xlsx, PHARMA.xlsx, "
            "and/or ORGANISER.xlsx in the project folder."
        )
    stats = sync_guests_from_rows(all_rows, replace_all=replace_all)
    stats["files"] = loaded_files
    return stats


@app.route("/reload-fixed-excels", methods=["POST"])
@admin_required
def reload_fixed_excels():
    """Re-import fixed category Excel files (keeps QR/ID for same name+phone)."""
    try:
        stats = load_fixed_guest_excels(replace_all=True)
        flash(
            f"Loaded {stats['total']} guest(s) from: "
            + "; ".join(stats.get("files") or [])
            + f". Kept same QR/ID: {stats['kept']}, new: {stats['created']}.",
            "ok",
        )
    except Exception as exc:  # noqa: BLE001
        flash(f"Excel reload failed: {exc}", "bad")
    return redirect(url_for("home"))


@app.route("/sample-excel")
@admin_required
def sample_excel():
    if SAMPLE_XLSX.exists():
        return send_file(
            SAMPLE_XLSX,
            as_attachment=True,
            download_name=SAMPLE_XLSX.name,
        )
    flash("Sample Excel file is missing on the server.", "bad")
    return redirect(url_for("home"))


@app.route("/ack/<token>")
def acknowledge(token: str):
    """Guest clicks from email — no staff login required."""
    guest = get_guest_by_token(token)
    if guest is None:
        return render_template("ack.html", guest=None, already=False), 404

    already = bool(guest["email_acked"])
    if not already:
        conn = get_db()
        conn.execute(
            """
            UPDATE guests
            SET email_acked = 1, email_acked_at = ?
            WHERE token = ?
            """,
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), token),
        )
        conn.commit()
        conn.close()
        guest = get_guest_by_token(token)

    return render_template("ack.html", guest=guest, already=already)


def _safe_pdf_filename(guest: sqlite3.Row) -> str:
    """Unique download name: Name_SHORTID_pass.pdf"""
    name = re.sub(r"[^\w\-]+", "_", (guest["name"] or "guest").strip(), flags=re.UNICODE)
    name = name.strip("_") or "guest"
    code = guest_badge_id(guest["token"])
    return f"{name}_{code}_pass.pdf"


@app.route("/pdf/<token>")
@login_required
def pdf_ticket(token: str):
    guest = get_guest_by_token(token)
    if guest is None:
        abort(404)
    pdf_buffer = build_ticket_pdf(guest)
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=_safe_pdf_filename(guest),
    )


@app.route("/download-pdfs-zip")
@login_required
def download_pdfs_zip():
    """One-click ZIP of all badge PDFs (optional ?category=Delegate)."""
    raw_cat = (request.args.get("category") or "").strip()
    category = normalize_category(raw_cat) if raw_cat else None
    if raw_cat and category not in CATEGORIES:
        category = None

    conn = get_db()
    if category:
        guests = conn.execute(
            "SELECT * FROM guests WHERE designation = ? ORDER BY name COLLATE NOCASE, id",
            (category,),
        ).fetchall()
    else:
        guests = conn.execute(
            "SELECT * FROM guests ORDER BY designation, name COLLATE NOCASE, id"
        ).fetchall()
    conn.close()

    if not guests:
        flash("No guests to download.", "bad")
        return redirect(url_for("home", category=category or None))

    zip_buffer = io.BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for guest in guests:
            pdf_buffer = build_ticket_pdf(guest)
            filename = _safe_pdf_filename(guest)
            # Folder by category inside the zip for easier printing
            cat = normalize_category(guest["designation"])
            arcname = f"{cat}/{filename}"
            if arcname in used_names:
                stem = filename[:-4] if filename.lower().endswith(".pdf") else filename
                arcname = f"{cat}/{stem}_{guest['id']}.pdf"
            used_names.add(arcname)
            zf.writestr(arcname, pdf_buffer.read())

    zip_buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    if category:
        zip_name = f"NGPA_{category}_badges_{stamp}.zip"
    else:
        zip_name = f"NGPA_all_badges_{stamp}.zip"
    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=zip_name,
    )


@app.route("/send/<int:guest_id>", methods=["POST"])
@login_required
def send_one(guest_id: int):
    try:
        guest = get_guest_by_id(guest_id)
        if guest is None:
            flash("Guest not found.", "bad")
            return redirect(url_for("home"))
        if not guest["email"] or "@" not in guest["email"]:
            flash(
                f"{guest['name']} has no email — download PDF instead, or add email in Excel.",
                "bad",
            )
            return redirect(url_for("home"))
        pdf_buffer = build_ticket_pdf(guest)
        send_pdf_email(guest, pdf_buffer)
        mark_pdf_sent(guest_id)
        flash(f"QR PDF sent to {guest['name']} ({guest['email']}).", "ok")
    except Exception as exc:  # noqa: BLE001 — show setup errors to staff
        flash(f"Could not send email: {exc}", "bad")
    return redirect(url_for("home"))


@app.route("/send-all", methods=["POST"])
@login_required
def send_all():
    try:
        conn = get_db()
        guests = conn.execute("SELECT * FROM guests ORDER BY id").fetchall()
        conn.close()

        ok_count = 0
        errors = []
        for guest in guests:
            try:
                if not guest["email"] or "@" not in guest["email"]:
                    errors.append(f"{guest['name']}: no email")
                    continue
                pdf_buffer = build_ticket_pdf(guest)
                send_pdf_email(guest, pdf_buffer)
                mark_pdf_sent(guest["id"])
                ok_count += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{guest['name']}: {exc}")

        if ok_count:
            flash(f"Sent QR PDFs to {ok_count} guest(s).", "ok")
        if errors:
            flash("Some failed - " + " | ".join(errors[:3]), "bad")
        if not ok_count and not errors:
            flash("No guests to email.", "bad")
    except Exception as exc:  # noqa: BLE001
        flash(f"Send-all failed: {exc}", "bad")
    return redirect(url_for("home"))


def _mark_arrived_if_needed(conn: sqlite3.Connection, guest: sqlite3.Row, token: str, now: str) -> None:
    if not guest["arrived"]:
        conn.execute(
            "UPDATE guests SET arrived = 1, arrived_at = ? WHERE token = ?",
            (now, token),
        )


@app.route("/checkin/<token>", methods=["GET", "POST"])
@login_required
def checkin(token: str):
    guest = get_guest_by_token(token)
    focus_meal = (request.args.get("meal") or request.form.get("meal") or "").lower()
    if focus_meal not in {"lunch", "dinner"}:
        focus_meal = ""

    if guest is None:
        return (
            render_template(
                "checkin.html", guest=None, status="invalid", focus_meal=focus_meal
            ),
            404,
        )

    message = None
    if request.method == "POST":
        action = request.form.get("action")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn = get_db()

        if action == "arrive":
            if guest["arrived"]:
                message = "Guest was already marked arrived."
            else:
                conn.execute(
                    "UPDATE guests SET arrived = 1, arrived_at = ? WHERE token = ?",
                    (now, token),
                )
                message = "Guest marked as arrived."
        elif action == "lunch":
            if guest["lunch_claimed"]:
                message = "Lunch was already claimed."
            else:
                _mark_arrived_if_needed(conn, guest, token, now)
                conn.execute(
                    "UPDATE guests SET lunch_claimed = 1, claimed_at = ? WHERE token = ?",
                    (now, token),
                )
                message = "Lunch marked as claimed."
        elif action == "dinner":
            if guest["dinner_claimed"]:
                message = "Dinner was already claimed."
            else:
                _mark_arrived_if_needed(conn, guest, token, now)
                conn.execute(
                    """
                    UPDATE guests
                    SET dinner_claimed = 1, dinner_claimed_at = ?
                    WHERE token = ?
                    """,
                    (now, token),
                )
                message = "Dinner marked as claimed."
        else:
            message = "Unknown action."

        conn.commit()
        conn.close()
        guest = get_guest_by_token(token)

    if guest["lunch_claimed"] and guest["dinner_claimed"]:
        status = "claimed"
    elif guest["lunch_claimed"] or guest["dinner_claimed"] or guest["arrived"]:
        status = "arrived"
    else:
        status = "ok"

    return render_template(
        "checkin.html",
        guest=guest,
        status=status,
        message=message,
        focus_meal=focus_meal,
    )


@app.route("/scan", methods=["GET", "POST"])
@login_required
def scan():
    if request.method == "POST":
        guest = resolve_guest_code(request.form.get("code", ""))
        if guest:
            return redirect(url_for("checkin", token=guest["token"]))
        flash(
            "Invalid QR / ID — this guest is not registered.",
            "error",
        )
        return redirect(url_for("scan"))
    return render_template("scan.html")


@app.route("/health")
def health():
    return {"status": "ok"}


# Runs for both `python app.py` and gunicorn (Render / production)
init_db()
try:
    _conn = get_db()
    _count = _conn.execute("SELECT COUNT(*) AS c FROM guests").fetchone()["c"]
    _conn.close()
    if _count == 0 and any(path.exists() for path, _cat in FIXED_GUEST_EXCELS):
        with app.app_context():
            load_fixed_guest_excels(replace_all=True)
except Exception:
    pass


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG") == "1")
